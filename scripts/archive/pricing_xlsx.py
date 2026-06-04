#!/usr/bin/env python3
"""Generate Model Pricing Excel file."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Model Pricing"

# --- Data ---
headers = ["MODEL", "INPUT", "OUTPUT", "CACHED READ", "CACHED WRITE"]

rows = [
    ["Big Pickle", "Free", "Free", "Free", "-"],
    ["DeepSeek V4 Flash Free", "Free", "Free", "Free", "-"],
    ["MiMo-V2.5 Free", "Free", "Free", "Free", "-"],
    ["Nemotron 3 Super Free", "Free", "Free", "Free", "-"],
    ["MiniMax M2.7", "$0.30", "$1.20", "$0.06", "$0.375"],
    ["MiniMax M2.5", "$0.30", "$1.20", "$0.06", "$0.375"],
    ["GLM 5.1", "$1.40", "$4.40", "$0.26", "-"],
    ["GLM 5", "$1.00", "$3.20", "$0.20", "-"],
    ["Kimi K2.5", "$0.60", "$3.00", "$0.10", "-"],
    ["Kimi K2.6", "$0.95", "$4.00", "$0.16", "-"],
    ["Qwen3.7 Max", "$2.50", "$7.50", "$0.50", "$3.125"],
    ["Qwen3.6 Plus", "$0.50", "$3.00", "$0.05", "$0.625"],
    ["Qwen3.5 Plus", "$0.20", "$1.20", "$0.02", "$0.25"],
    ["DeepSeek V4 Flash", "$0.14", "$0.28", "$0.03", "-"],
    ["Grok Build 0.1", "$1.00", "$2.00", "$0.20", "-"],
    ["Claude Opus 4.8", "$5.00", "$25.00", "$0.50", "$6.25"],
    ["Claude Opus 4.7", "$5.00", "$25.00", "$0.50", "$6.25"],
    ["Claude Opus 4.6", "$5.00", "$25.00", "$0.50", "$6.25"],
    ["Claude Opus 4.5", "$5.00", "$25.00", "$0.50", "$6.25"],
    ["Claude Opus 4.1", "$15.00", "$75.00", "$1.50", "$18.75"],
    ["Claude Sonnet 4.6", "$3.00", "$15.00", "$0.30", "$3.75"],
    ["Claude Sonnet 4.5 (≤ 200K tokens)", "$3.00", "$15.00", "$0.30", "$3.75"],
    ["Claude Sonnet 4.5 (> 200K tokens)", "$6.00", "$22.50", "$0.60", "$7.50"],
    ["Claude Sonnet 4 (≤ 200K tokens)", "$3.00", "$15.00", "$0.30", "$3.75"],
    ["Claude Sonnet 4 (> 200K tokens)", "$6.00", "$22.50", "$0.60", "$7.50"],
    ["Claude Haiku 4.5", "$1.00", "$5.00", "$0.10", "$1.25"],
    ["Gemini 3.5 Flash", "$1.50", "$9.00", "$0.15", "-"],
    ["Gemini 3.1 Pro (≤ 200K tokens)", "$2.00", "$12.00", "$0.20", "-"],
    ["Gemini 3.1 Pro (> 200K tokens)", "$4.00", "$18.00", "$0.40", "-"],
    ["Gemini 3 Flash", "$0.50", "$3.00", "$0.05", "-"],
    ["GPT 5.5 (≤ 272K tokens)", "$5.00", "$30.00", "$0.50", "-"],
    ["GPT 5.5 (> 272K tokens)", "$10.00", "$45.00", "$1.00", "-"],
    ["GPT 5.5 Pro", "$30.00", "$180.00", "$30.00", "-"],
    ["GPT 5.4 (≤ 272K tokens)", "$2.50", "$15.00", "$0.25", "-"],
    ["GPT 5.4 (> 272K tokens)", "$5.00", "$22.50", "$0.50", "-"],
    ["GPT 5.4 Pro", "$30.00", "$180.00", "$30.00", "-"],
    ["GPT 5.4 Mini", "$0.75", "$4.50", "$0.075", "-"],
    ["GPT 5.4 Nano", "$0.20", "$1.25", "$0.02", "-"],
    ["GPT 5.3 Codex Spark", "$1.75", "$14.00", "$0.175", "-"],
    ["GPT 5.3 Codex", "$1.75", "$14.00", "$0.175", "-"],
    ["GPT 5.2", "$1.75", "$14.00", "$0.175", "-"],
    ["GPT 5.2 Codex", "$1.75", "$14.00", "$0.175", "-"],
    ["GPT 5.1", "$1.07", "$8.50", "$0.107", "-"],
    ["GPT 5.1 Codex", "$1.07", "$8.50", "$0.107", "-"],
    ["GPT 5.1 Codex Max", "$1.25", "$10.00", "$0.125", "-"],
    ["GPT 5.1 Codex Mini", "$0.25", "$2.00", "$0.025", "-"],
    ["GPT 5", "$1.07", "$8.50", "$0.107", "-"],
    ["GPT 5 Codex", "$1.07", "$8.50", "$0.107", "-"],
    ["GPT 5 Nano", "$0.05", "$0.40", "$0.005", "-"],
]

# --- Styles ---
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
free_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
data_font = Font(name="Calibri", size=11)
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

# --- Title row ---
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = "Model Pricing \u2014 Pay-as-you-go (per 1M tokens)"
title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# --- Subtitle row ---
ws.merge_cells("A2:E2")
sub_cell = ws["A2"]
sub_cell.value = "Prices in USD. Claude Haiku 3.5 is used for session title generation only."
sub_cell.font = Font(name="Calibri", italic=True, size=9, color="808080")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# --- Headers (row 3) ---
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws.row_dimensions[3].height = 25

# --- Data rows ---
for row_idx, row_data in enumerate(rows, 4):
    is_free = row_data[1] == "Free"
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border
        if col_idx == 1:
            cell.alignment = left_align
        else:
            cell.alignment = data_align
        if is_free:
            cell.fill = free_fill

# --- Separator line after free models ---
for col in range(1, 6):
    ws.cell(row=7, column=col).border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="medium", color="2F5496"),
    )

# Column widths
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 16

# Freeze header rows
ws.freeze_panes = "A4"

# Auto-filter
ws.auto_filter.ref = f"A3:E{3 + len(rows)}"

output_path = "/Users/luke/Desktop/Model_Pricing.xlsx"
wb.save(output_path)
print(f"Done \u2014 saved to {output_path}")
print(f"Rows: {len(rows)} models")
